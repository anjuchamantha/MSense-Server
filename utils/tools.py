import os
import statistics
from datetime import datetime
from functools import reduce

import numpy as np
import openpyxl
import pandas as pd
from imblearn.over_sampling import SMOTE
from imblearn.under_sampling import RandomUnderSampler
from openpyxl.styles import PatternFill

import FeatureSelection
import io
from utils.general import colab


def top_users(df):
    return df['user_id'].value_counts()


def drop_na_rows_radius_of_gyration(df):
    return df[df['radius_of_gyration'].notna()]


def oversampling_using_smote(df, x_cols=io.x_cols, y_cols=io.y_col):
    x = df[x_cols]
    y = df[y_cols]
    y = y.astype('int')
    # count = y["meal_taken"].value_counts()
    # print('\n# of 1 = ', count.get(1))
    # print('# of 0 = ', count.get(0))

    oversample = SMOTE(k_neighbors=2)
    x, y = oversample.fit_resample(x, y)

    # count = y["meal_taken"].value_counts()
    # print('\n# of 1 = ', count.get(1))
    # print('# of 0 = ', count.get(0))

    # print('row count after oversampling = ', len(y.index))
    return [x, y]


def normalize(df):
    result = pd.DataFrame(columns=df.columns)
    for feature_name in df.loc[:, df.columns != 'user_id']:
        max_value = df[feature_name].max()
        min_value = df[feature_name].min()
        # print(feature_name,min_value,max_value)
        if max_value == min_value:
            result[feature_name] = float(max_value)
        else:
            result[feature_name] = (df[feature_name] - min_value) / (max_value - min_value)
    result['user_id'] = df['user_id']
    return result


# get the test data without the users in train set
def get_test_data(df, train_data):
    new_df = df.loc[~df['user_id'].isin(train_data["user_id"].unique())]
    test_data = pd.DataFrame(columns=train_data.columns)
    test_data = test_data.append(new_df, ignore_index=True)
    return test_data


def calculate_stat_values(arr):
    return [statistics.mean(arr), statistics.median(arr), statistics.stdev(arr)]


def calculate_stat_values_all(arr):
    return [statistics.mean(arr), statistics.stdev(arr), statistics.median(arr), min(arr), max(arr)]


# filter out the given user data from the dataframe
def df_without_selected_user(df, user):
    new_df = df.loc[~df['user_id'].isin([user])]

    df2 = pd.DataFrame(columns=new_df.columns)
    df2 = df2.append(new_df, ignore_index=True)

    for feature_name in df.loc[:, df.columns != 'user_id']:
        df2[feature_name] = df2[feature_name].astype(str).astype(float)

    df2['meal_taken'] = df2['meal_taken'].apply(np.int64)
    return df2


def print_mean_med_sd(clf, arr):
    data = calculate_stat_values(arr)
    print(clf, "\t =", data[0], data[1], data[2])
    return data


def print_mean_sd(clf, arr):
    data = calculate_stat_values(arr)
    print(clf, "\t =", data[0], data[2])
    return [data[0], data[2]]


def print_mean_sd_median_min_max(clf, arr):
    data = calculate_stat_values_all(arr)
    print(clf, "\t =", data[0], data[1], data[2], data[3], data[4])
    return [data[0], data[1], data[2], data[3], data[4]]


def reduce_dimension_excel(read_file):
    write_file = read_file + "-reduced"
    # write_file = read_file + "-sum"

    if colab:
        wbk_path = '/content/drive/MyDrive/FYP-Colab/' + write_file + ".xlsx"
        read_wbk_path = '/content/drive/MyDrive/FYP-Colab/' + read_file + ".xlsx"

    else:
        wbk_path = "../results/" + write_file + ".xlsx"
        read_wbk_path = "../results/" + read_file + ".xlsx"

    read_wbk = openpyxl.load_workbook(read_wbk_path)

    if os.path.isfile(wbk_path) and os.access(wbk_path, os.R_OK):
        print("[File] Writing to the existing file : " + wbk_path)
        wbk = openpyxl.load_workbook(wbk_path)
        wbk_sheet = wbk.create_sheet()
    else:
        print("[File] Making a new xlsx file")
        wbk = openpyxl.Workbook()
        wbk_sheet = wbk.active

    sheet1 = read_wbk.worksheets[0]

    # read and loop read_wbk
    for row_i in range(1, sheet1.max_row + 1):  # For each Row in a worksheet
        for column_i in range(1, sheet1.max_column + 1):
            if row_i != 1 and column_i != 1:
                vals = []
                for sheet in read_wbk.worksheets:  # For each worksheet
                    val = sheet.cell(row=row_i, column=column_i).value
                    if type(val) == str or val is None:
                        vals.append(0)
                    else:
                        vals.append(val)
                # print(vals)
                avg = statistics.mean(vals)
                # avg = sum(vals)
                wbk_sheet.cell(row=row_i, column=column_i).value = avg
                if column_i % 5 == 2:
                    wbk_sheet.cell(row=row_i, column=column_i).fill = PatternFill("solid", fgColor="00FFCC99")
            else:
                wbk_sheet.cell(row=row_i, column=column_i).value = sheet1.cell(row=row_i, column=column_i).value

    wbk.save(wbk_path)
    wbk.close()


def reduce_dimension_excel_f_groups(read_file, s_cell=11, e_cell=17):
    write_file = read_file + "-reduced-f-groups"

    if colab:
        wbk_path = '/content/drive/MyDrive/FYP-Colab/' + write_file + ".xlsx"
        read_wbk_path = '/content/drive/MyDrive/FYP-Colab/' + read_file + ".xlsx"

    else:
        wbk_path = "../results/" + write_file + ".xlsx"
        read_wbk_path = "../results/" + read_file + ".xlsx"

    read_wbk = openpyxl.load_workbook(read_wbk_path)

    if os.path.isfile(wbk_path) and os.access(wbk_path, os.R_OK):
        print("[File] Writing to the existing file : " + wbk_path)
        wbk = openpyxl.load_workbook(wbk_path)
        wbk_sheet = wbk.create_sheet()
    else:
        print("[File] Making a new xlsx file")
        wbk = openpyxl.Workbook()
        wbk_sheet = wbk.active

    sheet1 = read_wbk.worksheets[0]

    row_id = 2
    for sheet in read_wbk.worksheets:  # For each user
        print("sheet", sheet)
        wbk_sheet.cell(row=row_id, column=1).value = str(sheet)
        for column_i in range(1, sheet1.max_column + 1):  # for each feature
            print("col = ", column_i)
            wbk_sheet.cell(row=1, column=column_i).value = sheet1.cell(row=1, column=column_i).value
            col_vals = []
            for row_i in range(s_cell, e_cell + 1):  # For each Row in a worksheet
                if row_i != 1 and column_i != 1:
                    val = sheet.cell(row=row_i, column=column_i).value
                    col_vals.append(val)

            if len(col_vals) != 0:
                max_col_val = max(col_vals)
                print(max_col_val, col_vals)
                wbk_sheet.cell(row=row_id, column=column_i).value = max_col_val
                if column_i % 5 == 2:
                    wbk_sheet.cell(row=row_id, column=column_i).fill = PatternFill("solid", fgColor="00FFCC99")

        row_id += 1

    wbk.save(wbk_path)
    wbk.close()


def write_to_excel(arr, filename="results", sheet_name=None):
    if colab:
        wbk_path = '/content/drive/MyDrive/FYP-Colab/' + filename + ".xlsx"

    else:
        wbk_path = "E:/FYP/DataClassifiers/results/" + filename + ".xlsx"
    # print(arr)

    if os.path.isfile(wbk_path) and os.access(wbk_path, os.R_OK):
        print("[File] Writing to the existing file : " + wbk_path)
        wbk = openpyxl.load_workbook(wbk_path)
        if sheet_name is not None:
            wbk_sheet = wbk.create_sheet(sheet_name)
        else:
            wbk_sheet = wbk.create_sheet()
    else:
        print("[File] Making a new xlsx file")
        wbk = openpyxl.Workbook()
        wbk_sheet = wbk.active
        if sheet_name is not None:
            wbk_sheet.title = sheet_name

    for myRow in range(1, len(arr) + 1):  # 1 to 43
        for myCol in range(1, len(arr[myRow - 1]) + 1):  # 1 to 27
            wbk_sheet.cell(row=myRow, column=myCol).value = arr[myRow - 1][myCol - 1]
            if myCol % 5 == 2:
                wbk_sheet.cell(row=myRow, column=myCol).fill = PatternFill("solid", fgColor="00CCFFCC")
    wbk.save(wbk_path)
    wbk.close()


def impute_missing_data_using_mean(dataframe):
    # print('null values in =', dataframe.columns[dataframe.isna().any()].tolist())
    print("Impute Missing Data Using Mean in columns - ", dataframe.columns[dataframe.isna().any()].tolist())

    columns = dataframe.columns[dataframe.isna().any()].tolist()

    for col in columns:
        dataframe[col] = dataframe[col].fillna((dataframe[col].mean()))
        # print(dataframe[col].head(10))

    return dataframe


def calculate_stat_values_of_confusion_matrix_values(arr):
    output = []
    for score in range(len(arr[0]) - 4):
        values = []
        for iteration in range(len(arr)):
            values.append(arr[iteration][score])
        output.append(calculate_stat_values_all(values))

    # for tn,fp,tp,fn scores
    for score in range(len(arr[0]) - 4, len(arr[0])):
        values = []
        for iteration in range(len(arr)):
            values.append(arr[iteration][score])
        output.append([sum(values)])

    return output


def print_mean_sd_median_min_max_of_confusion_matrix_values(clf, arr):
    results = calculate_stat_values_of_confusion_matrix_values(arr)
    flatten_arr = reduce(lambda z, y: z + y, results)

    print(clf, "\t =", flatten_arr)

    return flatten_arr + ["-"]


def print_model_results(rf_list, nb_list, gb_list, mlp_keras_list, svm_list, xgb_list, adab_list):
    print("\t\t\t\t\t\t\t\t\t\t Accuracy \t\t\t\t\t\t\t\t\t\t\t"
          "\t\t\t\t\t\t\t\t\t\t F1-Score \t\t\t\t\t\t\t\t\t\t\t"
          "\t\t\t\t\t\t\t\t\t\t AUC-ROC \t\t\t\t\t\t\t\t\t\t\t"
          "\t\t\t\t\t\t\t\t\t\t Kappa \t\t\t\t\t\t\t\t\t\t\t")
    print("\t\t\t avg \t\t\t\t sd \t\t\t\t median \t\t\t\t min \t\t\t\t max ||"
          "\t\t\t avg \t\t\t\t sd \t\t\t\t median \t\t\t\t min \t\t\t\t max ||"
          "\t\t\t avg \t\t\t\t sd \t\t\t\t median \t\t\t\t min \t\t\t\t max ||"
          "\t\t\t avg \t\t\t\t sd \t\t\t\t median \t\t\t\t min \t\t\t\t max")

    rf_data = print_mean_sd_median_min_max_of_confusion_matrix_values("RF ", rf_list)
    nb_data = print_mean_sd_median_min_max_of_confusion_matrix_values("NB ", nb_list)
    gb_data = print_mean_sd_median_min_max_of_confusion_matrix_values("GB ", gb_list)
    mlp_k_data = print_mean_sd_median_min_max_of_confusion_matrix_values("MLP_K", mlp_keras_list)
    svm_data = print_mean_sd_median_min_max_of_confusion_matrix_values("SVM", svm_list)
    xgb_data = print_mean_sd_median_min_max_of_confusion_matrix_values("XGB", xgb_list)
    adab_data = print_mean_sd_median_min_max_of_confusion_matrix_values("AdaB", adab_list)

    return rf_data + nb_data + gb_data + mlp_k_data + svm_data + xgb_data + adab_data


def find_eating_noteating(dataframe):
    return dataframe['meal_taken'].value_counts()


def weekend_weekday_label_encoding(dataframe):
    # pd.set_option('display.max_columns', None)
    for index, row in dataframe.iterrows():
        d = datetime(2019, row['month'], row['date'])

        if d.weekday() > 4:  # weekend = 1
            dataframe.loc[index, 'weekend'] = 1
            # print(d, d.weekday(), ' date is weekend.')
        else:  # weekday = 0
            dataframe.loc[index, 'weekend'] = 0
            # print(d, d.weekday(), ' data is weekday.')

    return dataframe


def feature_selection(dataframe, x_columns, y_columns, filename):
    feature_arr = []
    n_features = len(x_columns)
    while n_features > 5:
        n_features -= 5
        x_columns = FeatureSelection.SequentialForwardFeatureSelection(dataframe, x_columns, y_columns,
                                                                       n_features=n_features)
        feature_arr.append(x_columns)

    filename = "../results/" + filename + ".txt"
    with open(filename, "w") as text_file:
        for grp in feature_arr:
            text_file.write("%s \n" % grp)
    return feature_arr


def read_from_txt(filename):
    if colab:
        file = open("/content/drive/MyDrive/FYP-Colab/FYP/results/" + filename + ".txt", "r")

    else:
        file = open("../results/" + filename + ".txt", "r")

    groups = []
    lines = file.read().splitlines()
    for line in lines:
        # print(line)
        x_ = line.strip('\'\'[ ]').replace("'", '').replace(" ", "")
        # print(x_.replace(" ", ""))
        x = x_.split(',')
        # print(x)
        # print()
        groups.append(x)
    # print("Txt Read :")
    # print(groups)
    return groups


def under_sampling_2_class(df, x_cols=io.x_cols, y_col=io.y_col):
    x = df[x_cols]
    y = df[y_col]
    y = y.astype('int')

    undersample = RandomUnderSampler()
    x, y = undersample.fit_resample(x, y)

    return [x, y]


def random_over_under_sampling_2_class(df, x_cols=io.x_cols, y_col=io.y_col):
    x = df[x_cols]
    y = df[y_col]
    y = y.astype('int')

    # print("random sampling")
    # print("before sampling = ", len(x.index))
    # print(y[y_col].value_counts())

    value_counts = y[y_col].value_counts()
    i = 0
    for index, value in value_counts.items():
        if i == 0 and index[0] == 0:  # most no of rows with = 0
            oversample = SMOTE(sampling_strategy={0: value_counts.get(0),
                                                  1: min(value_counts.get(0), value_counts.get(1) * 2)},
                               k_neighbors=1)
            x, y = oversample.fit_resample(x, y)
            i += 1
        elif i == 0 and index[0] == 1:  # most no of rows with = 1
            oversample = SMOTE(sampling_strategy={0: min(value_counts.get(0) * 2, value_counts.get(1)),
                                                  1: value_counts.get(1)},
                               k_neighbors=1)
            x, y = oversample.fit_resample(x, y)
            i += 1

    # print("\nafter oversampling = ", len(x.index))
    # print(y[y_col].value_counts())

    undersample = RandomUnderSampler()
    # rus = RandomUnderSampler(random_state=0, ratio={0: 30, 1: 20, 2: 60})
    # undersample.fit(x, y)
    x, y = undersample.fit_resample(x, y)

    # print("\nafter under sampling = ", len(x.index))
    # print(y[y_col].value_counts())
    # plt.hist(y[y_col])
    # plt.show()

    return [x, y]


def write_user_selection_to_excel():
    excel_sheet = []
    excel_first_row = ["F Groups"]
    excel_sheet.append(excel_first_row)
    feature_selection_arr = read_from_txt(filename="Experiment3/10_af576877ade8f9a2aca82655bedcac1896cdef87")
    for grp_num in range(len(feature_selection_arr)):
        group_name = "U1 F" + str(grp_num + 1)
        grp_val = feature_selection_arr[grp_num]
        print(group_name, grp_val)
        excel_sheet.append([group_name] + grp_val)
    write_to_excel(excel_sheet, filename="Experiment3/Exp3_Users_feature_selection")
